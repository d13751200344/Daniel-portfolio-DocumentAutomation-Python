import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os
from openpyxl.styles import Font, PatternFill
import time


def columnIntToLetter(index):
    # Convert a column index (1-based) to its corresponding letter (e.g., 1 -> 'A', 2 -> 'B', etc.)
    # Convert the index to 0-based by subtracting 1
    index -= 1
    # Calculate the quotient and remainder
    quotient = index // 26
    remainder = index % 26
    # Convert the remainder to ASCII and add 65 to get the corresponding letter
    letter = chr(65 + remainder)
    # If the quotient is greater than 0, recursively call the function and append the letter
    if quotient > 0:
        letter = columnIntToLetter(quotient) + letter

    return letter


def columnStrToInt(input):
    if not input.isalpha():
        raise ValueError(
            "Input must be a string containing only alphabetical characters.")

    input = input.upper()
    result = 0
    for char in input:
        result = (ord(char) - 64) + result * 26
    return result


def findProject(input):
    # Iterate through all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and (str(cell.value).strip().lower() == input.strip().lower()):
                result = [columnIntToLetter(cell.column), cell.row]
                # Return the row and column of the cell as column letter and row number
                return result
                # [ columnString, row ]

    # If the value is not found, return None
    return None


def findColumnStartRowEndRow(input, sheet):
    cellLocation = findProject(input)
    if cellLocation is not None:
        # print(f'Column is {cellLocation[0]}, row is {cellLocation[1]}.')
        # ['A', 4]
        startRow = cellLocation[1] + 1
        endRowCount = 4
        row = 4
        while sheet.cell(row=row, column=1).value is not None:
            row += 1
            endRowCount += 1
        endRow = endRowCount-1
        return [cellLocation[0], startRow, endRow]
        # [the begin column, begin row, end row] = ['A', 5, 784]

# Get begin & last column of faculty & student


def findStartAndEndColumn(input):
    n = columnStrToInt(findProject(input)[0])
    findLast = False
    while not findLast:
        cellValue = sheet.cell(row=4, column=n).value.strip().lower()
        if cellValue is None or not (input[0:-1] in cellValue):
            findLast = True
        else:
            n += 1
    # print(n-1)
    return [columnStrToInt(findProject(input)[0]), n-1]


def msgWindow(msg):
    # create a pop-up window
    window = tk.Toplevel()
    window.title("Message")
    window.geometry("300x300")
    label = tk.Label(window, text=msg, wraplength=250)
    label.pack(pady=20)
    # add a "Close" button that will destroy the window when clicked
    closeBtn = tk.Button(window, text="OK", command=window.quit)
    closeBtn.pack(pady=10)
    window.mainloop()


# Welcome window:
root = tk.Tk()
root.title("Welcome")
root.geometry("300x100")
label = tk.Label(root, text="Welcome! Please select a file in next step.")
label.pack(pady=10)


def btnClick():
    print("Click OK and begin to select a file...")
    global correctFile
    correctFile = False
    root.quit()
    root.destroy()


button = tk.Button(root, text="OK", command=btnClick)
button.pack()


def shut():
    root.quit()
    root.destroy()


# if click 'x', shut down the program
root.protocol("WM_DELETE_WINDOW", shut)
root.mainloop()


while not correctFile:
    try:
        # get the absolute path of the .xlsx file
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename()

        if not file_path:
            # User clicked 'Cancel', exit the loop
            break

        if file_path and file_path[-4:] == 'xlsx':
            abs_path = os.path.abspath(file_path)
            print("File path:", abs_path)

            class ExcelLoader:
                def __init__(self, master, abs_path):
                    self.master = master
                    self.master.title("Excel Loader")
                    self.abs_path = abs_path

                    # Create labels
                    self.loading_label = tk.Label(
                        self.master, text="Loading may take you a few seconds, please wait...", font=("Arial", 15), wraplength=250)
                    self.loading_label.pack(pady=30)

                    self.error_label = tk.Label(
                        self.master, text="", font=("Arial", 15), fg="red", wraplength=250)
                    self.error_label.pack(pady=10)

                    # Call the open_excel_file method
                    self.open_excel_file()

                def open_excel_file(self):
                    self.master.update()
                    try:
                        global workbook
                        workbook = openpyxl.load_workbook(self.abs_path)
                        if workbook:
                            root.destroy()
                            root.quit()
                    except:
                        self.error_label.config(text="Error opening file!")

            # Create main window
            root = tk.Tk()
            root.geometry("400x150")
            app = ExcelLoader(root, abs_path)
            # root.mainloop()

            if workbook:
                correctFile = True

        else:
            msg = "Invalid file. Please select a valid .xlsx file."
            print(msg)
            messagebox.showinfo(
                "Warning", msg)

    except Exception:
        msg = "Unexpected Error."
        messagebox.showinfo(
            "Warning", msg)
        print(msg)

sheet = workbook.active

haveAllColumn = False

if (findProject('project #') is None) or (findProject('Activity Fiscal Year') is None) or (findProject('project status') is None) or (findProject('Company') is None) or (findProject('faculty #1') is None) or (findProject('student #1') is None):
    msgWindow("Please make sure the selecting excel file has columns named 'Project #', 'Activity Fiscal Year', 'project status', 'Company', 'faculty #1', and 'student #1' and restart this program again.")
else:
    projectNumList = findColumnStartRowEndRow('Project #', sheet)
    # ['A', 5, 784]

    # project #, company, faculty, student, ... (4)
    columnNameRow = projectNumList[1]-1
    startRow = projectNumList[1]  # where contents begin (5)
    lastRow = projectNumList[2]  # the last row of project# (784)

    # get columns of fiscal year, project#, faculty names, student names, company name, and status
    fiscalYearColumn = columnStrToInt(findProject('Activity Fiscal Year')[
                                      0])  # the column of 'Activity Fiscal Year'
    statusColumn = columnStrToInt(findProject('project status')[
                                  0])  # the column of 'Project Status'
    projectColumn = columnStrToInt(
        projectNumList[0])  # the column of 'Project #'
    companyColumn = columnStrToInt(findProject(
        'Company')[0])  # the column of 'Company'

    # get begin & last column of faculty [24, 28]
    findStartAndEndColumn('faculty #1')

    # get begin & last column of student [31, 50]
    findStartAndEndColumn('student #1')

    # ask fiscal year > if enter a not existing value, keep asking
    fiscalYearOption = []
    for i in range(startRow, lastRow+1):
        if (sheet.cell(row=i, column=fiscalYearColumn).value is not None) and (sheet.cell(row=i, column=fiscalYearColumn).value not in fiscalYearOption):
            fiscalYearOption.append(sheet.cell(
                row=i, column=fiscalYearColumn).value)

    haveAllColumn = True


while haveAllColumn:
    option = tk.Tk()
    option.title("Fiscal Year")
    option.geometry("300x300")
    label = tk.Label(option, text="Select a fiscal year:")
    label.pack()
    var = tk.StringVar(option)
    var.set(None)  # make radio buttons not pre-selected

    for year in fiscalYearOption:
        yearOption = tk.Radiobutton(
            option, text=year, variable=var, value=year)
        yearOption.pack()

    def buttonClick():
        global selectedYear
        selectedYear = var.get()
        print("Selected fiscal year:", selectedYear)
        option.quit()
        option.destroy()

    button = tk.Button(option, text="Select", command=buttonClick)
    button.pack()

    # if click 'x', shut down the program
    option.protocol("WM_DELETE_WINDOW", shut)  # use shut() that defines before
    option.mainloop()

    projectCountSet = set()  # count the number of projects
    companyCountSet = set()  # count the number of companies
    facultyCountSet = set()  # count the number of faculties
    studentCountSet = set()  # count the number of students
    dataList = []
    count = int(startRow)

    while count <= int(lastRow):
        if (sheet.cell(row=count, column=fiscalYearColumn).value == str(selectedYear)):
            if (sheet.cell(row=count, column=statusColumn).value.strip().lower() == 'completed') or (sheet.cell(row=count, column=statusColumn).value.strip().lower() == 'in progress') or (sheet.cell(row=count, column=statusColumn).value[:9].strip().lower() == 'duplicate') or (sheet.cell(row=count, column=statusColumn).value[:16].strip().lower() == 'waiting for sign'):

                # count the number of projects
                projectCountSet.add(sheet.cell(
                    row=count, column=projectColumn).value)
                # count the number of companies
                companyCountSet.add(sheet.cell(
                    row=count, column=companyColumn).value)

                rowList = []
                rowList.append(sheet.cell(
                    row=count, column=projectColumn).value)
                rowList.append(sheet.cell(
                    row=count, column=companyColumn).value)

                facultyList = []
                for i in range(findStartAndEndColumn("faculty #1")[0], (findStartAndEndColumn("faculty #1")[1] + 1)):
                    if sheet.cell(row=count, column=i).value is not None:
                        facultyList.append(sheet.cell(
                            row=count, column=i).value)
                        # count the number of faculties
                        facultyCountSet.add(sheet.cell(
                            row=count, column=i).value)

                studentList = []
                for j in range(findStartAndEndColumn("student #1")[0], (findStartAndEndColumn("student #1")[1] + 1)):
                    if sheet.cell(row=count, column=j).value is not None:
                        studentList.append(sheet.cell(
                            row=count, column=j).value)
                        # count the number of students
                        studentCountSet.add(sheet.cell(
                            row=count, column=j).value)
                rowList.append(facultyList)
                rowList.append(studentList)
                dataList.append(rowList)
        count += 1

    projectNumbers = len(projectCountSet)
    companyNumbers = len(companyCountSet)
    facultyNumbers = len(facultyCountSet)
    studentNumbers = len(studentCountSet)

    print(f'''We have {projectNumbers} projects: {projectCountSet}, 
    {companyNumbers} companies: {companyCountSet}; 
    {facultyNumbers} faculties: {facultyCountSet}; 
    and {studentNumbers} students: {studentCountSet} in the specified year.''')

    print(f'Detailed data: {dataList}')
    ''' [['GEO 727', 'Bcompany', ['danise', 'emily'], ['Ff', 'Gg', 'Ee']], 
    ['GEO 728', 'Ccompany', ['frank', 'emily'], ['Ii', 'Ee', 'Hh']]] '''

    wb = openpyxl.Workbook()
    sheetOne = wb.active
    sheetOne.title = 'total'
    sheetOne['B2'] = "Metrics"
    sheetOne['B3'] = "Projects"
    sheetOne['B4'] = "Companies"
    sheetOne['B5'] = "Faculty Researchers"
    sheetOne['B6'] = "Student Researchers"

    sheetOne['C2'] = selectedYear
    sheetOne['C3'] = projectNumbers
    sheetOne['C4'] = companyNumbers
    sheetOne['C5'] = facultyNumbers
    sheetOne['C6'] = studentNumbers

    sheetTwo = wb.create_sheet(title='details')

    # Style the sheet
    sheetOne.column_dimensions['B'].width = 20
    sheetOne.column_dimensions['C'].width = 20
    sheetTwo.column_dimensions['B'].width = 20
    sheetTwo.column_dimensions['C'].width = 20
    greyFill = PatternFill(
        fill_type="solid", start_color='808080', end_color='808080')
    lightGreyFill = PatternFill(
        fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    blueFill = PatternFill(
        fill_type="solid", start_color="4472C4", end_color="4472C4")
    lightBlueFill = PatternFill(
        fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
    # sheet["B2"].fill = blueFill
    columnList = ['B', 'C']
    for i in columnList:
        cell = i+'2'
        sheetOne[cell].fill = greyFill
        sheetOne[cell].font = Font(bold=True, color='FFFFFF')
    for i in columnList:
        for j in range(3, 7):
            cell = i+str(j)
            sheetOne[cell].fill = blueFill
            sheetOne[cell].font = Font(bold=True, color='FFFFFF')

    columnNameList = ['Project #', 'Company', 'Faculty #1', 'Faculty #2', 'Faculty #3', 'Faculty #4', 'Faculty #5', 'Student #1', 'Student #2', 'Student #3', 'Student #4', 'Student #5', 'Student #6',
                      'Student #7', 'Student #8', 'Student #9', 'Student #10', 'Student #11', 'Student #12', 'Student #13', 'Student #14', 'Student #15', 'Student #16', 'Student #17', 'Student #18', 'Student #19', 'Student #20']
    for i in range(len(columnNameList)):
        sheetTwo[columnIntToLetter(i+2) + str(2)] = columnNameList[i]

    # project # = B3 start
    # company # = C3 start
    for i in range(len(dataList)):
        sheetTwo.cell(row=i+3, column=2, value=dataList[i][0])
        sheetTwo.cell(row=i+3, column=3, value=dataList[i][1])

    # faculty # = D3 start, H3 last
    for i in range(len(dataList)):
        for j in range(4, 9):
            if len(dataList[i][2]) >= j-3 and dataList[i][2][j-4]:
                sheetTwo.cell(row=i+3, column=j, value=dataList[i][2][j-4])
            else:
                sheetTwo.cell(row=i+3, column=j, value=None)

    # student # = I3 start, AB3 last
    for i in range(len(dataList)):
        for j in range(9, 29):
            if len(dataList[i][3]) >= j-8 and dataList[i][3][j-9]:
                sheetTwo.cell(row=i+3, column=j, value=dataList[i][3][j-9])
            else:
                sheetTwo.cell(row=i+3, column=j, value=None)

    root = tk.Tk()
    root.withdraw()  # Hide the tkinter window
    savePath = filedialog.asksaveasfilename()

    if not savePath:
        # User clicked 'Cancel', exit the loop
        break

    # Save the file to the chosen location
    if savePath:
        wb.save(savePath + '.xlsx')
        wb.close()
        messagebox.showinfo(
            "Congrats", "Congratulations! Your file is complete.")
        haveAllColumn = False
